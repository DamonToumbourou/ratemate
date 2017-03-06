from flask import Flask, render_template, url_for, g, redirect, session, request, flash
from scrapers import WebScrapers
from openpyxl import Workbook
import private
import os
import sqlite3
from datetime import datetime, timedelta
import time

app = Flask(__name__)
app.config.from_object(__name__) 

app.config.update(dict(
    DATABASE=os.path.join(app.root_path, 'ratemate.db'),
    SECRET_KEY=private.secret_key,
    USERNAME='admin',
    PASSSWORD=private.password
))
app.config.from_envvar('FLASKR_SETTINGS', silent=True)


"""
Database handlers
"""

def connect_db():
    rv = sqlite3.connect(app.config['DATABASE'])
    rv.row_factory = sqlite3.Row
    return rv

def get_db():
    if not hasattr(g, 'sqlite_db'):
        g.sqlite_db = connect_db()
    return g.sqlite_db

def init_db():
    db = get_db()
    with app.open_resource('schema.sql', mode='r') as f:
        db.cursor().executescript(f.read())
    db.commit()

@app.cli.command('initdb')
def initdb_command():
    init_db()
    print('init the database')


@app.teardown_appcontext
def close_db(error):
    if hasattr(g, 'sqlite_db'):
        g.sqlite_db.close()

"""
Routes for pages
"""
@app.route('/') 
def home():
    
    return render_template('index.html')


@app.route('/term_deposit')
def term_deposit():
    db = get_db()
    
    try:
        most_recent_time = db.execute('SELECT MAX(date) from term_deposit')
        most_recent_time = most_recent_time.fetchall()
        most_recent_time = most_recent_time[0][0] + 120 
        time_range = most_recent_time - 300 

        cur = db.execute('select logo, date, name, product, one_month, two_month, three_month, ' \
                         'four_month, five_month, six_month, seven_month, eight_month, nine_month, ' \
                         'ten_month, eleven_month, twelve_month, twentyfour_month, thirtysix_month, '\
                         'date from term_deposit WHERE date BETWEEN ' + str(time_range) + ' AND ' + str(most_recent_time) )
        term_deposit = cur.fetchall()
    
    except:
        term_deposit = None
        print 'no term deposits found'

    """ 
    highest =  get_highest_td()
    big_4_td = get_big_4_td()
    """

    return render_template('term_deposit.html', term_deposit=term_deposit)


@app.route('/online_saver')
def online_saver():
    db = get_db()
    
    try:
        most_recent_time = db.execute('SELECT MAX(date) from online_saver')
        most_recent_time = most_recent_time.fetchall()
        most_recent_time = most_recent_time[0][0] + 120 
        time_range = most_recent_time - 300 

        cur = db.execute('select logo, date, name, notes, product, base, bonus, total, ' \
                         'date from online_saver WHERE date BETWEEN ' + str(time_range) + ' AND ' + str(most_recent_time) )
        online_saver = cur.fetchall()
    
    except:
        online_saver = None
        print 'no online savers found'


    return render_template('online_saver.html', online_saver=online_saver)


@app.route('/progress_saver')
def progress_saver():
    db = get_db()
    
    try:
        most_recent_time = db.execute('SELECT MAX(date) from progress_saver')
        most_recent_time = most_recent_time.fetchall()
        most_recent_time = most_recent_time[0][0] + 120 
        time_range = most_recent_time - 300 

        cur = db.execute('select logo, date, name, notes, product, base, bonus, total, ' \
                         'date from progress_saver WHERE date BETWEEN ' + str(time_range) + ' AND ' + str(most_recent_time) )
        progress_saver = cur.fetchall()
    
    except:
        progress_saver = None
        print 'no progress savers found'

    return render_template('progress_saver.html', progress_saver=progress_saver)


@app.route('/cash_manager')
def cash_manager():
    db = get_db()
    
    try:
        most_recent_time = db.execute('SELECT MAX(date) from cash_manager')
        most_recent_time = most_recent_time.fetchall()
        most_recent_time = most_recent_time[0][0] + 120 
        time_range = most_recent_time - 300 

        cur = db.execute('select logo, date, name, notes, product, _500k, _250k, _100k, _50k, ' \
                         'date from cash_manager WHERE date BETWEEN ' + str(time_range) + ' AND ' + str(most_recent_time) )
        cash_manager = cur.fetchall()
    
    except:
        cash_manager = None
        print 'no cash manager found'
    print cash_manager
    return render_template('cash_manager.html', cash_manager=cash_manager)


@app.route('/pensioner')
def pensioner():
    db = get_db()
    
    try:
        most_recent_time = db.execute('SELECT MAX(date) from pensioner')
        most_recent_time = most_recent_time.fetchall()
        most_recent_time = most_recent_time[0][0] + 120 
        time_range = most_recent_time - 300 

        cur = db.execute('select logo, date, name, notes, product, _0k, _2k, _5K, ' \
                         'date from pensioner WHERE date BETWEEN ' + str(time_range) + ' AND ' + str(most_recent_time) )
        pensioner = cur.fetchall()
    
    except:
        pensioner = None
        print 'no online pensioners found'

    return render_template('pensioner.html', pensioner=pensioner)


"""
Routes for scraping data
"""

@app.route('/add_online')
def add_online():
    db = get_db()
    db = db.execute('select CAST(date AS float) from online_saver')
    dates = db.fetchall()
    uptodate = False
    
    # if online db is empty
    if not dates:
        print 'DB empty... fetching Online Savers'
        write_online_db()

    else:
        # if DB has entries then check if has been scraped in the last 24hours
        seconds_week = 604800
        now = time.time()
        time_past = now - seconds_week

        for date in dates:
            if time_past < date[0]:
                uptodate = True
                date_scraped = time.strftime('%d-%m-%Y %H:%M', time.localtime(date[0]))
    
        if not uptodate:
            try:
                write_online_db()
                flash('Database updated...')
            except:
                flash('Database update failed')
    
    if uptodate:
        flash('Database is up to date. Last scraped: ')
        flash(date_scraped)


    return redirect(url_for('online_saver'))


@app.route('/add_progress')
def add_progress():
    db = get_db()
    db = db.execute('select CAST(date AS float) from progress_saver')
    dates = db.fetchall()
    uptodate = False
    
    # if online db is empty
    if not dates:
        print 'DB empty... fetching Progress Savers'
        write_progress_db()

    else:
        # if DB has entries then check if has been scraped in the last 24hours
        seconds_week = 604800
        now = time.time()
        time_past = now - seconds_week

        for date in dates:
            if time_past < date[0]:
                uptodate = True
                date_scraped = time.strftime('%d-%m-%Y %H:%M', time.localtime(date[0]))
    
        if not uptodate:
            try:
                write_progress_db()
                flash('Database updated...')
            except:
                flash('Database update failed')
    
    if uptodate:
        flash('Database is up to date. Last scraped: ')
        flash(date_scraped)


    return redirect(url_for('progress_saver'))


@app.route('/add_cash')
def add_cash():
    db = get_db()
    db = db.execute('select CAST(date AS float) from cash_manager')
    dates = db.fetchall()
    uptodate = False
    
    # if online db is empty
    if not dates:
        print 'DB empty... fetching cash managers'
        write_cash_db()

    else:
        # if DB has entries then check if has been scraped in the last 24hours
        seconds_week = 604800
        now = time.time()
        time_past = now - seconds_week

        for date in dates:
            if time_past < date[0]:
                uptodate = True
                date_scraped = time.strftime('%d-%m-%Y %H:%M', time.localtime(date[0]))
    
        if not uptodate:
            try:
                write_cash_db()
                flash('Database updated...')
            except:
                flash('Database update failed')
    
    if uptodate:
        flash('Database is up to date. Last scraped: ')
        flash(date_scraped)


    return redirect(url_for('cash_manager'))


@app.route('/add_pensioner')
def add_pensioner():
    db = get_db()
    db = db.execute('select CAST(date AS float) from pensioner')
    dates = db.fetchall()
    uptodate = False
    
    # if online db is empty
    if not dates:
        print 'DB empty... fetching pensioner'
        write_pensioner_db()

    else:
        # if DB has entries then check if has been scraped in the last 24hours
        seconds_week = 604800
        now = time.time()
        time_past = now - seconds_week

        for date in dates:
            if time_past < date[0]:
                uptodate = True
                date_scraped = time.strftime('%d-%m-%Y %H:%M', time.localtime(date[0]))
    
        if not uptodate:
            try:
                write_pensioner_db()
                flash('Database updated...')
            except:
                flash('Database update failed')
    
    if uptodate:
        flash('Database is up to date. Last scraped: ')
        flash(date_scraped)

    return redirect(url_for('pensioner'))


@app.route('/add_td', methods=['GET'])
def add_td():
    
    db = get_db()
    db = db.execute('select CAST(date AS float) from term_deposit')
    dates = db.fetchall()
    uptodate = False
    
    # if td db is empty
    if not dates:
        print 'DB empty... fetching TDs'
        write_td_db()

    else:
        # if DB has entries then check if has been scraped in the last 24hours
        seconds_week = 604800
        now = time.time()
        time_past = now - seconds_week

        for date in dates:
            if time_past < date[0]:
                uptodate = True
                date_scraped = time.strftime('%d-%m-%Y %H:%M', time.localtime(date[0]))
    
        if not uptodate:
            try:
                write_td_db()
                flash('Database updated...')
            except:
                flash('Database update failed')
    
    if uptodate:
        flash('Database is up to date. Last scraped: ')
        flash(date_scraped)

    return redirect(url_for('term_deposit'))


"""
Functions for handling data
"""

def write_td_db():
    results = WebScrapers()
    results = results.collate_td()
    db_write = get_db()
    
    for result in results: 
        name = result[1].get('name')
        product = result[1].get('product')
        logo = result[1].get('logo')
        date = time.time()
        one_month = result[0].get('one_month')
        two_month = result[0].get('two_month')
        three_month = result[0].get('three_month')
        four_month = result[0].get('four_month')
        five_month = result[0].get('five_month')
        six_month = result[0].get('six_month')
        seven_month = result[0].get('seven_month')
        eight_month = result[0].get('eight_month')
        nine_month = result[0].get('nine_month')
        ten_month = result[0].get('ten_month')
        eleven_month = result[0].get('eleven_month')
        twelve_month = result[0].get('twelve_month')
        twentyfour_month = result[0].get('twentyfour_month')
        thirtysix_month = result[0].get('thirtysix_month')
        
        db_write.execute('insert into term_deposit (name, logo, product, one_month, two_month, ' \
                         'three_month, four_month, five_month, six_month, seven_month, eight_month, ' \
                         'nine_month, ten_month, eleven_month, twelve_month, twentyfour_month, ' \
                         'thirtysix_month, date) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', \
                         [name, logo, product, one_month, two_month, three_month, four_month, \
                          five_month, six_month, seven_month, eight_month, nine_month, ten_month, \
                          eleven_month, twelve_month, twentyfour_month, thirtysix_month, date])
        
        db_write.commit()
    
    return True


def write_online_db():
    results = WebScrapers()
    results = results.collate_online_savers()
    db_write = get_db()
    
    for result in results: 
        name = result[1].get('name')
        product = result[1].get('product')
        logo = result[1].get('logo')
        date = time.time()
        base = result[0].get('base')
        bonus = result[0].get('bonus')
        total = result[0].get('total')
        notes = result[1].get('notes')

        db_write.execute('insert into online_saver (name, logo, product, notes, base, bonus, total, ' \
                         'date) values (?, ?, ?, ?, ?, ?, ?, ?)', \
                         [name, logo, product, notes, base, bonus, total, date])
        
        db_write.commit()
    
    return True


def write_progress_db():
    results = WebScrapers()
    results = results.collate_progress_saver()
    db_write = get_db()
    
    for result in results: 
        name = result[1].get('name')
        product = result[1].get('product')
        logo = result[1].get('logo')
        date = time.time()
        base = result[0].get('base')
        bonus = result[0].get('bonus')
        total = result[0].get('total')
        notes = result[1].get('notes')

        db_write.execute('insert into progress_saver (name, logo, product, notes, base, bonus, total, ' \
                         'date) values (?, ?, ?, ?, ?, ?, ?, ?)', \
                         [name, logo, product, notes, base, bonus, total, date])
        
        db_write.commit()
    
    return True


def write_pensioner_db():
    results = WebScrapers()
    results = results.collate_pensioner()
    db_write = get_db()
    
    for result in results: 
        name = result[1].get('name')
        product = result[1].get('product')
        logo = result[1].get('logo')
        date = time.time()
        _0k = result[0].get('0k')
        _2k = result[0].get('2k')
        _5k = result[0].get('5k')
        notes = result[1].get('notes')
        print '_0k'
        print _0k
        db_write.execute('insert into pensioner (name, logo, product, notes, _0k, _2k, _5k, date) values (?, ?, ?, ?, ?, ?, ?, ?)', [name, logo, product, notes, _0k, _2k, _5k, date])
        
        db_write.commit()
    
    return True


def write_cash_db():
    results = WebScrapers()
    results = results.collate_cash()
    db_write = get_db()
    
    for result in results: 
        name = result[1].get('name')
        product = result[1].get('product')
        logo = result[1].get('logo')
        date = time.time()
        _500k = result[0].get('500k and over')
        _250k = result[0].get('250k - 500k')
        _100k = result[0].get('100k - 250k')
        _50k = result[0].get('50k - 100k')
        notes = result[1].get('notes')

        db_write.execute('insert into cash_manager (name, logo, product, notes, _500k, _250k, _100k, _50k, ' \
                         'date) values (?, ?, ?, ?, ?, ?, ?, ?, ?)', \
                         [name, logo, product, notes, _500k, _250k, _100k, _50k, date])
        
        db_write.commit()
    
    return True


def get_highest_td():
    db = get_db()
    
    latest_date = db.execute('SELECT MAX(date) from term_deposit')
    latest_date = latest_date.fetchall()
    latest_date = latest_date[0][0] + 60
    range_date = latest_date - 300

    db = db.execute('SELECT name, logo, short, short_rate, mid, mid_rate, long, long_rate, date FROM term_deposit WHERE date BETWEEN ' + str(range_date) + ' AND ' + str(latest_date) )
    terms = db.fetchall()
    
    short = 0
    mid = 0
    _long = 0 
    highest = []
    for term in terms:
        
        cur_short = term['short_rate'].strip('%')
        if short < cur_short: 
            short_name = term['name']
            short = cur_short
            short_logo = term['logo']
            short_days = term['short']
            
        cur_mid = term['mid_rate'].strip('%')
        if mid < cur_mid:
            mid_name = term['name']
            mid = cur_mid
            mid_logo = term['logo']
            mid_days = term['mid']

        cur_long = term['long_rate'].strip('%')
        if _long < cur_long:
            long_name = term['name']
            _long = cur_long
            long_logo = term['logo']
            long_days = term['long']

    highest.append([{
        'short_rate': short,
        'short_days': short_days,
        'short_name': short_name,
        'short_logo': short_logo,
        'mid_rate': mid,
        'mid_days': mid_days,
        'mid_name': mid_name,
        'mid_logo': mid_logo,
        'long_rate': _long,
        'long_name': long_name,
        'long_logo': long_logo
    }])
    
    return highest


def get_big_4_td():
    db = get_db()
    
    latest_date = db.execute('SELECT MAX(date) from term_deposit')
    latest_date = latest_date.fetchall()
    latest_date = latest_date[0][0] + 60
    range_date = latest_date - 300


    db = db.execute('''SELECT name, logo, short, short_rate, mid, mid_rate, long, long_rate, date FROM term_deposit WHERE( date BETWEEN ''' + str(range_date) + ''' AND ''' + str(latest_date) + ''' ) AND ( name LIKE '%CBA%' OR name LIKE '%westpac%' OR name LIKE '%NAB%' OR name LIKE '%ANZ Advanced%' )''')
    
    terms = db.fetchall()
    
    big_4 = []
    for term in terms: 
        big_4.append([term]) 
   
    return big_4


"""
functions for making excel workbooks
"""
def write_td_workbook():
    wb = Workbook()
    ws = wb.active
     
    count = 1
    col_2_2nd = 'A'
    col_2 = 'C'
    col_3_2nd = 'B'
    col_3 = 'D'
    new_count = 0
    new_count_3 = 1
    for term in term_deposit:
        ws['A2'] = term[1] # date 
        ws['B2'] = 'Short term' 
        ws['B3'] = 'Mid term'
        ws['B4'] = 'Long term'
        ws[ col_2 + '1'] = term[2] # bank name
        ws[ col_2 + '2'] = int(term[3]) # days short
        ws[ col_2 + '3'] = int(term[5]) # days mid 
        ws[ col_2 + '4'] = int(term[7]) # days long

        if term[4]:
            ws[ col_3 + '2'] = float(term[4].strip('%')) # rate short
            ws[ col_3 + '3'] = float(term[6].strip('%')) # rate short
            ws[ col_3 + '4'] = float(term[8].strip('%')) # rate short

        count = count + 3
         
        if count <= 23:
            col_2 = chr(ord(col_2) + 3)
            col_3 = chr(ord(col_3) + 3)

        if count >= 24:
            col_2 = 'A' + chr(ord(col_2_2nd) + (new_count))
            col_2 = 'A' + chr(ord(col_3_2nd) + (new_count_3))
            
            new_count = new_count + 3

    wb.save('static/test.xlsx')
    
    highest =  get_highest_td()
    big_4_td = get_big_4_td()
    
    return None


if __name__ == "__main__":
    app.run()
